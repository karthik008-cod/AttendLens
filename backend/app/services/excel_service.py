import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.db.models import Classroom, Student, LectureDate, AttendanceRecord

REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

def generate_class_excel(db, classroom_id: int) -> str:
    classroom = db.classrooms.find_one({"id": classroom_id})
    if not classroom:
        raise ValueError("Classroom not found")

    students = list(db.students.find({"classroom_id": classroom_id}).sort("roll_number", 1))
    lecture_dates = list(db.lecture_dates.find({"classroom_id": classroom_id}).sort("date_str", 1))
    records = list(db.attendance_records.find({"classroom_id": classroom_id}))

    # Create lookup dictionary: (student_id, lecture_date_id) -> status
    status_map = {(r["student_id"], r["lecture_date_id"]): r["status"] for r in records}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{classroom['name'][:30]} Attendance"

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    normal_font = Font(name="Segoe UI", size=10)
    alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alert_font = Font(name="Segoe UI", size=10, bold=True, color="9C0006")
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    good_font = Font(name="Segoe UI", size=10, color="006100")
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Headers (built first so we know the total column count)
    headers = ["Student ID / Roll No", "Student Name", "Total %"] + [ld["date_str"] for ld in lecture_dates]

    # Title Block — merge spans ALL columns dynamically
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = f"Classroom: {classroom['name']} ({classroom['subject']})"
    ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
    ws["A1"].alignment = left_align

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for idx, student in enumerate(students):
        row_idx = 4 + idx
        roll_cell = ws.cell(row=row_idx, column=1, value=student["roll_number"])
        roll_cell.alignment = center_align
        roll_cell.font = bold_font
        roll_cell.border = thin_border

        name_cell = ws.cell(row=row_idx, column=2, value=student["name"])
        name_cell.alignment = left_align
        name_cell.font = bold_font
        name_cell.border = thin_border

        # Calculate attendance and fill date columns starting at column 4
        present_count = 0
        for col_offset, ld in enumerate(lecture_dates):
            status = status_map.get((student["id"], ld["id"]), "-")
            if status == "P":
                present_count += 1
            cell = ws.cell(row=row_idx, column=4 + col_offset, value=status)
            cell.alignment = center_align
            cell.border = thin_border
            if status == "P":
                cell.font = Font(name="Segoe UI", size=10, color="008000")
            elif status == "A":
                cell.font = Font(name="Segoe UI", size=10, color="C00000", bold=True)
            else:
                cell.font = normal_font

        # Column 3: Total %
        pct = round((present_count / total_lectures * 100), 1) if total_lectures > 0 else 0.0
        pct_cell = ws.cell(row=row_idx, column=3, value=f"{pct}%")
        pct_cell.alignment = center_align
        pct_cell.border = thin_border
        if total_lectures > 0 and pct < 75.0:
            pct_cell.fill = alert_fill
            pct_cell.font = alert_font
        else:
            pct_cell.fill = good_fill
            pct_cell.font = good_font

    # Auto-fit columns (default=10 guards against empty columns)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    file_path = os.path.join(REPORTS_DIR, f"Attendance_{classroom.name.replace(' ', '_')}_{classroom.id}.xlsx")
    wb.save(file_path)
    return file_path
